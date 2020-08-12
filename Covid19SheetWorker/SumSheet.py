import openpyxl as xl
wb = xl.load_workbook("29July2020.xlsx")
source_sheet = wb['Effects-of-COVID-19-on-trade-1-']
target_sheet = wb['ValueByTransport']
t_row = 1

s_column_namelist = []
for column in range(1, source_sheet.max_column + 1):
    s_column_namelist.append(source_sheet.cell(1, column).value)
    print(s_column_namelist)

# get column index of certain value
